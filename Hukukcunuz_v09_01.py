from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
from datetime import datetime, timedelta
import os
import openpyxl
from openpyxl.utils import get_column_letter
import re

# Botun token'ı (Telegram BotFather'dan almanız gerekiyor)
TOKEN = "BOT API BURAYA"

# Mevcut tarih ve saat (Türkiye saat dilimi +03:00)
CURRENT_TIME = "19 Haziran 2025, 08:30"

# Kullanıcı dil tercihleri, son mesaj ID'leri ve randevu durumu
USER_LANGUAGE = {}
USER_LAST_MESSAGE = {}
USER_APPOINTMENT_STATE = {}  # Randevu alma sürecini takip etmek için

# Excel dosya adı
APPOINTMENT_FILE = "appointments.xlsx"

# Excel dosyasını oluştur veya yükle
def init_appointment_excel():
    if not os.path.exists(APPOINTMENT_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Randevular"
        headers = ["Tarih", "Saat", "Kullanıcı ID", "Kullanıcı Adı", "Müşteri Adı", "Telefon", "Detaylar", "Durum"]
        ws.append(headers)

        # 10 günlük randevu slotlarını oluştur
        start_date = datetime.now()
        hours = ["09:00", "10:00", "11:00", "12:00", "13:00", "14:00", "15:00", "16:00", "17:00"]
        for day in range(10):
            date = (start_date + timedelta(days=day)).strftime("%d.%m.%Y")
            for hour in hours:
                ws.append([date, hour, "", "", "", "", "", "Boş"])
        wb.save(APPOINTMENT_FILE)

# Randevu slotlarını oku
def get_available_slots(date=None):
    wb = openpyxl.load_workbook(APPOINTMENT_FILE)
    ws = wb.active
    slots = []
    current_date = datetime.now().strftime("%d.%m.%Y")
    for row in ws.iter_rows(min_row=2, values_only=True):
        slot_date, slot_time, _, _, _, _, _, status = row
        if status == "Boş" and (date is None or slot_date == date) and slot_date >= current_date:
            slots.append((slot_date, slot_time))
    wb.close()
    return slots

# Randevu kaydet
def save_appointment(user_id, username, name, phone, date, time, details):
    wb = openpyxl.load_workbook(APPOINTMENT_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == date and row[1].value == time and row[7].value == "Boş":
            row[2].value = user_id
            row[3].value = username
            row[4].value = name
            row[5].value = phone
            row[6].value = details
            row[7].value = "Dolu"
            break
    wb.save(APPOINTMENT_FILE)

# Telefon numarası doğrulama
def is_valid_phone(phone):
    pattern = r"^\+?\d{10,15}$"
    return bool(re.match(pattern, phone))

# Geri bildirimi dosyaya kaydet
def save_feedback(user_id, username, feedback):
    with open('feedback.txt', 'a', encoding='utf-8') as f:
        f.write(f"{CURRENT_TIME} | User ID: {user_id} | Username: {username} | Feedback: {feedback}\n")

# Ana menüyü oluştur
def get_main_menu(lang='tr'):
    keyboard = [
        [
            InlineKeyboardButton("📖 Hakkımızda" if lang == 'tr' else "📖 About Us", callback_data='about'),
            InlineKeyboardButton("🛠️ Hizmetler" if lang == 'tr' else "🛠️ Services", callback_data='services'),
        ],
        [
            InlineKeyboardButton("📞 İletişim" if lang == 'tr' else "📞 Contact", callback_data='contact'),
            InlineKeyboardButton("❓ SSS" if lang == 'tr' else "❓ FAQ", callback_data='faq'),
        ],
        [
            InlineKeyboardButton("📚 Hukuki Terimler" if lang == 'tr' else "📚 Legal Terms", callback_data='terms'),
            InlineKeyboardButton("💡 Hukuki İpuçları" if lang == 'tr' else "💡 Legal Tips", callback_data='tips'),
        ],
        [
            InlineKeyboardButton("🆘 Yardım" if lang == 'tr' else "🆘 Help", callback_data='help'),
            InlineKeyboardButton("🌐 Dil Değiştir" if lang == 'tr' else "🌐 Change Language", callback_data='language'),
        ],
        [
            InlineKeyboardButton("📝 Geri Bildirim" if lang == 'tr' else "📝 Feedback", callback_data='feedback'),
            InlineKeyboardButton("📅 Randevu Talebi" if lang == 'tr' else "📅 Appointment Request", callback_data='appointment'),
        ],
    ]
    return InlineKeyboardMarkup(keyboard)

# Geri butonlu menü oluştur
def get_back_button(lang='tr'):
    if lang == 'tr':
        return InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Geri", callback_data='back')]])
    else:
        return InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data='back')]])

# SSS alt menüsünü oluştur
def get_faq_menu(lang='tr'):
    keyboard = [
        [InlineKeyboardButton("💔 Boşanma Davası" if lang == 'tr' else "💔 Divorce Case", callback_data='faq_divorce')],
        [InlineKeyboardButton("🧾 Miras Hukuku" if lang == 'tr' else "🧾 Inheritance Law", callback_data='faq_inheritance')],
        [InlineKeyboardButton("📜 Kira Sözleşmesi" if lang == 'tr' else "📜 Lease Agreement", callback_data='faq_contract')],
        [InlineKeyboardButton("👮 Ceza Davası" if lang == 'tr' else "👮 Criminal Case", callback_data='faq_criminal')],
        [InlineKeyboardButton("💼 İş Hukuku" if lang == 'tr' else "💼 Labor Law", callback_data='faq_labor')],
        [InlineKeyboardButton("🛒 Tüketici Hakları" if lang == 'tr' else "🛒 Consumer Rights", callback_data='faq_consumer')],
        [InlineKeyboardButton("🏢 Ticari Hukuk" if lang == 'tr' else "🏢 Commercial Law", callback_data='faq_commercial')],
        [InlineKeyboardButton("💸 İcra Hukuku" if lang == 'tr' else "💸 Execution Law", callback_data='faq_execution')],
        [InlineKeyboardButton("⬅️ Geri" if lang == 'tr' else "⬅️ Back", callback_data='back')],
    ]
    return InlineKeyboardMarkup(keyboard)

# Hukuki Terimler alt menüsünü oluştur
def get_terms_menu(lang='tr'):
    keyboard = [
        [InlineKeyboardButton("📝 Vekaletname" if lang == 'tr' else "📝 Power of Attorney", callback_data='terms_proxy')],
        [InlineKeyboardButton("⏳ Dava Zamanaşımı" if lang == 'tr' else "⏳ Statute of Limitations", callback_data='terms_statute')],
        [InlineKeyboardButton("📜 Vasiyetname" if lang == 'tr' else "📜 Will", callback_data='terms_testament')],
        [InlineKeyboardButton("🚨 İhtiyati Tedbir" if lang == 'tr' else "🚨 Injunction", callback_data='terms_injunction')],
        [InlineKeyboardButton("📜 Sözleşme" if lang == 'tr' else "📜 Contract", callback_data='terms_contract_term')],
        [InlineKeyboardButton("💸 Haciz" if lang == 'tr' else "💸 Seizure", callback_data='terms_seizure')],
        [InlineKeyboardButton("⬅️ Geri" if lang == 'tr' else "⬅️ Back", callback_data='back')],
    ]
    return InlineKeyboardMarkup(keyboard)

# Hukuki İpuçları alt menüsünü oluştur
def get_tips_menu(lang='tr'):
    keyboard = [
        [InlineKeyboardButton("📝 Sözleşme İmzalamak" if lang == 'tr' else "📝 Signing a Contract", callback_data='tips_contract_sign')],
        [InlineKeyboardButton("🏛️ Mahkemeye Hazırlık" if lang == 'tr' else "🏛️ Court Preparation", callback_data='tips_court_preparation')],
        [InlineKeyboardButton("🛒 Tüketici Şikayeti" if lang == 'tr' else "🛒 Consumer Complaint", callback_data='tips_consumer_complaint')],
        [InlineKeyboardButton("📅 Randevu Planlama" if lang == 'tr' else "📅 Appointment Planning", callback_data='tips_appointment')],
        [InlineKeyboardButton("⬅️ Geri" if lang == 'tr' else "⬅️ Back", callback_data='back')],
    ]
    return InlineKeyboardMarkup(keyboard)

# Randevu tarih seçim menüsü
def get_date_menu(lang='tr'):
    slots = get_available_slots()
    dates = sorted(set(slot[0] for slot in slots))  # Benzersiz tarihler
    keyboard = [[InlineKeyboardButton(date, callback_data=f"date_{date}")] for date in dates[:10]]  # Maksimum 10 gün
    keyboard.append([InlineKeyboardButton("⬅️ Geri" if lang == 'tr' else "⬅️ Back", callback_data='back')])
    return InlineKeyboardMarkup(keyboard)

# Randevu saat seçim menüsü
def get_time_menu(date, lang='tr'):
    slots = get_available_slots(date)
    times = [slot[1] for slot in slots]
    keyboard = [[InlineKeyboardButton(time, callback_data=f"time_{date}_{time}")] for time in times]
    keyboard.append([InlineKeyboardButton("⬅️ Geri" if lang == 'tr' else "⬅️ Back", callback_data='appointment')])
    return InlineKeyboardMarkup(keyboard)

# Hukuk SSS verileri
FAQ_DATA = {
    'divorce': {
        'tr': "💔 **Boşanma Davası Nasıl Açılır?** ⚖️\nBoşanma davası, yetkili Aile Mahkemesinde dilekçe ile açılır. Gerekli belgeler: kimlik fotokopisi, evlenme cüzdanı ve varsa deliller (örneğin, tanık beyanları). Avukat tutmanız önerilir.\n📅 Son güncelleme: {}",
        'en': "💔 **How to File for Divorce?** ⚖️\nA divorce case is filed with a petition at the competent Family Court. Required documents: ID copy, marriage certificate, and any evidence (e.g., witness statements). Hiring a lawyer is recommended.\n📅 Last updated: {}"
    },
    'inheritance': {
        'tr': "🧾 **Miras Hukuku Nedir?** ⚖️\nMiras hukuku, vefat eden bir kişinin malvarlığının nasıl paylaşılacağını düzenler. Yasal mirasçılar (eş, çocuklar) önceliklidir; vasiyetname varsa atanmış mirasçılar da pay alabilir.\n📅 Son güncelleme: {}",
        'en': "🧾 **What is Inheritance Law?** ⚖️\nInheritance law regulates how a deceased person's assets are distributed. Legal heirs (spouse, children) have priority; a will can designate appointed heirs.\n📅 Last updated: {}"
    },
    'contract': {
        'tr': "📜 **Kira Sözleşmesi İptali Nasıl Yapılır?** ⚖️\nKira sözleşmesi, noter aracılığıyla ihtarname gönderilerek veya mahkeme yoluyla iptal edilebilir. Kiracı veya ev sahibi, sözleşmedeki şartlara uymazsa iptal talep edilebilir.\n📅 Son güncelleme: {}",
        'en': "📜 **How to Cancel a Lease Agreement?** ⚖️\nA lease agreement can be canceled via a notary notice or through a court. Cancellation can be requested if the tenant or landlord breaches contract terms.\n📅 Last updated: {}"
    },
    'criminal': {
        'tr': "👮 **Ceza Davası Nedir?** ⚖️\nCeza davası, bir suç isnadıyla açılan mahkeme sürecidir. Savcı iddianame hazırlar, sanık savunma yapar. Avukat desteği önemlidir.\n📅 Son güncelleme: {}",
        'en': "👮 **What is a Criminal Case?** ⚖️\nA criminal case is a court process initiated due to a crime allegation. The prosecutor prepares an indictment, and the defendant presents a defense. Legal representation is crucial.\n📅 Last updated: {}"
    },
    'labor': {
        'tr': "💼 **İş Hukuku Nedir?** ⚖️\nİş hukuku, işçi ve işveren arasındaki ilişkileri düzenler. Kıdem tazminatı, işten çıkarma ve çalışma koşulları gibi konuları kapsar.\n📅 Son güncelleme: {}",
        'en': "💼 **What is Labor Law?** ⚖️\nLabor law regulates relationships between employees and employers, covering issues like severance pay, termination, and working conditions.\n📅 Last updated: {}"
    },
    'consumer': {
        'tr': "🛒 **Tüketici Hakları Nedir?** ⚖️\nTüketici hakları, mal veya hizmet satın alan kişilerin korunmasını sağlar. Ayıplı mal için iade, değişim veya onarım talep edilebilir.\n📅 Son güncelleme: {}",
        'en': "🛒 **What are Consumer Rights?** ⚖️\nConsumer rights protect individuals purchasing goods or services. You can request a refund, exchange, or repair for defective products.\n📅 Last updated: {}"
    },
    'commercial': {
        'tr': "🏢 **Ticari Hukuk Nedir?** ⚖️\nTicari hukuk, şirketler ve ticari işlemlerle ilgili kuralları düzenler. Şirket kuruluşu, sözleşmeler ve ticari uyuşmazlıklar bu kapsamdadır.\n📅 Son güncelleme: {}",
        'en': "🏢 **What is Commercial Law?** ⚖️\nCommercial law regulates rules for companies and business transactions, covering company formation, contracts, and commercial disputes.\n📅 Last updated: {}"
    },
    'execution': {
        'tr': "💸 **İcra Hukuku Nedir?** ⚖️\nİcra hukuku, borçların tahsili ve alacaklıların haklarının korunması için uygulanan yasal süreçleri düzenler. İcra takibi noter veya icra dairesi aracılığıyla başlatılır.\n📅 Son güncelleme: {}",
        'en': "💸 **What is Execution Law?** ⚖️\nExecution law regulates legal processes for debt collection and protecting creditors' rights. Execution proceedings are initiated via a notary or execution office.\n📅 Last updated: {}"
    }
}

# Hukuki Terimler Sözlüğü
LEGAL_TERMS = {
    'proxy': {
        'tr': "📝 **Vekaletname Nedir?** ⚖️\nVekaletname, bir kişinin başka birine belirli bir işlem için yetki verdiği belgedir. Noter onayı gereklidir.\n📅 Son güncelleme: {}",
        'en': "📝 **What is a Power of Attorney?** ⚖️\nA power of attorney is a document authorizing someone to act on behalf of another in specific matters. Notary approval is required.\n📅 Last updated: {}"
    },
    'statute': {
        'tr': "⏳ **Dava Zamanaşımı Nedir?** ⚖️\nDava zamanaşımı, bir davanın açılabileceği yasal süreyi ifade eder. Örneğin, alacak davalarında genellikle 5 yıldır.\n📅 Son güncelleme: {}",
        'en': "⏳ **What is the Statute of Limitations?** ⚖️\nThe statute of limitations refers to the legal time limit for filing a case. For example, it is typically 5 years for debt claims.\n📅 Last updated: {}"
    },
    'testament': {
        'tr': "📜 **Vasiyetname Nedir?** ⚖️\nVasiyetname, bir kişinin vefatından sonra malvarlığının nasıl paylaşılacağını belirttiği belgedir. Noterde veya el yazısıyla hazırlanabilir.\n📅 Son güncelleme: {}",
        'en': "📜 **What is a Will?** ⚖️\nA will is a document specifying how a person's assets should be distributed after their death. It can be prepared at a notary or handwritten.\n📅 Last updated: {}"
    },
    'injunction': {
        'tr': "🚨 **İhtiyati Tedbir Nedir?** ⚖️\nİhtiyati tedbir, dava sürecinde hak kaybını önlemek için mahkemeden talep edilen geçici koruma önlemidir.\n📅 Son güncelleme: {}",
        'en': "🚨 **What is an Injunction?** ⚖️\nAn injunction is a temporary protective measure requested from the court to prevent loss of rights during a case.\n📅 Last updated: {}"
    },
    'contract_term': {
        'tr': "📜 **Sözleşme Nedir?** ⚖️\nSözleşme, iki veya daha fazla taraf arasında hak ve yükümlülükleri düzenleyen yasal bir anlaşmadır. Yazılı veya sözlü olabilir.\n📅 Son güncelleme: {}",
        'en': "📜 **What is a Contract?** ⚖️\nA contract is a legal agreement between two or more parties that outlines rights and obligations. It can be written or oral.\n📅 Last updated: {}"
    },
    'seizure': {
        'tr': "💸 **Haciz Nedir?** ⚖️\nHaciz, bir borçlunun malvarlığına devlet tarafından el konulması işlemidir. İcra müdürlüğü aracılığıyla yapılır.\n📅 Son güncelleme: {}",
        'en': "💸 **What is Seizure?** ⚖️\nSeizure is the process of state authorities seizing a debtor's assets, conducted through the execution office.\n📅 Last updated: {}"
    }
}

# Hukuki İpuçları
LEGAL_TIPS = {
    'contract_sign': {
        'tr': "📝 **Sözleşme İmzalamadan Önce Nelere Dikkat Edilmeli能够在？** ⚖️\n1. Sözleşmeyi dikkatlice okuyun.\n2. Tüm şartları anlayın ve belirsiz maddeleri sorun.\n3. İmzalamadan önce bir avukata danışın.\n📅 Son güncelleme: {}",
        'en': "📝 **What to Check Before Signing a Contract?** ⚖️\n1. Read the contract carefully.\n2. Understand all terms and clarify ambiguous clauses.\n3. Consult a lawyer before signing.\n📅 Last updated: {}"
    },
    'court_preparation': {
        'tr': "🏛️ **Mahkemeye Hazırlık İçin İpuçları** ⚖️\n1. Gerekli belgeleri toplayın.\n2. Avukatınızla strateji belirleyin.\n3. Mahkeme kurallarına uygun davranın.\n📅 Son güncelleme: {}",
        'en': "🏛️ **Tips for Preparing for Court** ⚖️\n1. Gather all necessary documents.\n2. Plan a strategy with your lawyer.\n3. Follow court rules and etiquette.\n📅 Last updated: {}"
    },
    'consumer_complaint': {
        'tr': "🛒 **Tüketici Şikayeti Nasıl Yapılır?** ⚖️\n1. Satıcıyla iletişime geçin.\n2. Sorun çözülmezse Tüketici Hakem Heyeti’ne başvurun.\n3. Gerekirse avukatla dava açın.\n📅 Son güncelleme: {}",
        'en': "🛒 **How to File a Consumer Complaint?** ⚖️\n1. Contact the seller.\n2. If unresolved, apply to the Consumer Arbitration Committee.\n3. If necessary, file a lawsuit with a lawyer.\n📅 Last updated: {}"
    },
    'appointment': {
        'tr': "📅 **Randevu Planlama İpuçları** ⚖️\n1. Randevu tarih ve saatini önceden kontrol edin.\n2. Konuyu netleştirin (örneğin, miras davası).\n3. Gerekli belgeleri yanınızda getirin.\n4. Doğru iletişim bilgileri (telefon numarası gibi) paylaşın.\n📅 Son güncelleme: {}",
        'en': "📅 **Appointment Planning Tips** ⚖️\n1. Check the date and time in advance.\n2. Clarify the topic (e.g., inheritance case).\n3. Bring necessary documents.\n4. Share accurate contact information (e.g., phone number).\n📅 Last updated: {}"
    }
}

# Başlangıç komutu
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    USER_LANGUAGE[user_id] = 'tr'  # Varsayılan dil Türkçe
    USER_APPOINTMENT_STATE[user_id] = None  # Randevu durumunu sıfırla
    init_appointment_excel()  # Excel dosyasını başlat

    welcome_message = (
        f"Merhaba! 🤗 ⚖️ LegalMind Hukuk Bürosuna hoş geldiniz.\n\n Akılcı çözümler, dijital hızla! 💼🤖 ⚖️🎉\n"
        f"🏛️ Hukuki sorularınız ve randevu talepleriniz için buradayız!\n"
        f"LegalMind Hukuk Bürosu, bireysel ve kurumsal müvekkillerine akılcı ve hızlı çözümler sunar ⚖️💼\n"
        f"Tüm işlemleriniz artık dijital asistanımız LegalMind Bot ile size bir tık kadar yakın! 🤖📲\n"
        f"📅 Son güncelleme: {CURRENT_TIME}\n"
        f"📋 Lütfen aşağıdaki menüden bir seçenek seçin:"
    )
    message = await update.message.reply_text(welcome_message, reply_markup=get_main_menu('tr'))
    USER_LAST_MESSAGE[user_id] = message.message_id

# Yardım komutu
async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    lang = USER_LANGUAGE.get(user_id, 'tr')
    USER_APPOINTMENT_STATE[user_id] = None  # Randevu durumunu sıfırla
    if lang == 'tr':
        message = (
            f"🆘 Yardım Menüsü:\n"
            f"➡️ /start: Botu başlatır ve ana menüyü gösterir.\n"
            f"➡️ Menüdeki butonlara tıklayarak bilgi alabilirsiniz.\n"
            f"➡️ Hukuki terimler, ipuçları ve randevu talebi için ilgili menüleri kullanın.\n"
            f"➡️ Geri bildirim bırakmak için 'Geri Bildirim' butonunu kullanın.\n"
            f"📅 Son güncelleme: {CURRENT_TIME}"
        )
    else:
        message = (
            f"🆘 Help Menu:\n"
            f"➡️ /start: Starts the bot and shows the main menu.\n"
            f"➡️ Click the menu buttons to get information.\n"
            f"➡️ Use Legal Terms, Legal Tips, and Appointment Request menus.\n"
            f"➡️ Use the 'Feedback' button to leave feedback.\n"
            f"📅 Last updated: {CURRENT_TIME}"
        )
    sent_message = await update.message.reply_text(message, reply_markup=get_back_button(lang))
    USER_LAST_MESSAGE[user_id] = sent_message.message_id

# Bilinmeyen komutlar için hata mesajı
async def unknown_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    lang = USER_LANGUAGE.get(user_id, 'tr')
    USER_APPOINTMENT_STATE[user_id] = None  # Randevu durumunu sıfırla
    if lang == 'tr':
        message = (
            f"😕 Üzgünüm, bu komutu anlayamadım.\n"
            f"➡️ /start veya /help ile devam edebilirsiniz!\n"
            f"📅 Son güncelleme: {CURRENT_TIME}"
        )
    else:
        message = (
            f"😕 Sorry, I didn't understand that command.\n"
            f"➡️ Try /start or /help!\n"
            f"📅 Last updated: {CURRENT_TIME}"
        )
    sent_message = await update.message.reply_text(message, reply_markup=get_back_button(lang))
    USER_LAST_MESSAGE[user_id] = sent_message.message_id

# Buton tıklama olaylarını işleme
async def button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    lang = USER_LANGUAGE.get(user_id, 'tr')

    # Mevcut mesaj içeriğini al
    current_text = query.message.text if query.message.text else ""
    current_reply_markup = query.message.reply_markup

    if query.data == 'back':
        USER_APPOINTMENT_STATE[user_id] = None
        message = (
            f"🏠 Ana menüye döndünüz! ⚖️\n"
            f"📅 Son güncelleme: {CURRENT_TIME}\n"
            f"📋 Lütfen aşağıdaki menüden bir seçenek seçin:"
        ) if lang == 'tr' else (
            f"🏠 Returned to the main menu! ⚖️\n"
            f"📅 Last updated: {CURRENT_TIME}\n"
            f"📋 Please select an option from the menu below:"
        )
        reply_markup = get_main_menu(lang)
    elif query.data == 'about':
        USER_APPOINTMENT_STATE[user_id] = None
        message = (
            f"📖 Hakkımızda:\n"
            f"Legalmind Hukuk hizmetleri olarak, kullanıcıları hukuki konularda bilgilendirir ve yönlendiririz! ⚖️\n"
            f"🏛️ Amacımız, hukuki süreçlerde size rehberlik etmek ve randevu taleplerinizi kolaylaştırmak.\n"
            f"📅 Son güncelleme: {CURRENT_TIME}"
        ) if lang == 'tr' else (
            f"📖 About Us:\n"
            f"We are Hukukçunuz Bot, guiding and informing users on legal matters! ⚖️\n"
            f"🏛️ Our goal is to assist you in legal processes and facilitate appointment requests.\n"
            f"📅 Last updated: {CURRENT_TIME}"
        )
        reply_markup = get_back_button(lang)
    elif query.data == 'services':
        USER_APPOINTMENT_STATE[user_id] = None
        message = (
            f"🛠️ Hizmetler:\n"
            f"✅ Hukuki danışmanlık ve rehberlik\n"
            f"✅ Hukuki terimler ve ipuçları\n"
            f"✅ Randevu talebi ve iletişim desteği\n"
            f"📅 Son güncelleme: {CURRENT_TIME}"
        ) if lang == 'tr' else (
            f"🛠️ Services:\n"
            f"✅ Legal consultancy and guidance\n"
            f"✅ Legal terms and tips\n"
            f"✅ Appointment requests and contact support\n"
            f"📅 Last updated: {CURRENT_TIME}"
        )
        reply_markup = get_back_button(lang)
    elif query.data == 'contact':
        USER_APPOINTMENT_STATE[user_id] = None
        message = (
            f"📞 İletişim Bilgileri:\n"
            f"📧 E-posta: yazilimci.men@gmail.com\n"
            f"📱 Telefon: +90 507 707 17 53\n"
            f"🌐 Web: https://luxury-souffle-655b07.netlify.app/\n"
            f"📅 Son güncelleme: {CURRENT_TIME}"
        ) if lang == 'tr' else (
            f"📞 Contact Information:\n"
            f"📧 Email: yazilimci.men@gmail.com\n"
            f"📱 Phone: +90 507 707 17 53\n"
            f"🌐 Website: https://luxury-souffle-655b07.netlify.app/\n"
            f"📅 Last updated: {CURRENT_TIME}"
        )
        reply_markup = get_back_button(lang)
    elif query.data == 'faq':
        USER_APPOINTMENT_STATE[user_id] = None
        message = (
            f"❓ Sıkça Sorulan Sorular:\n"
            f"📚 Lütfen bir hukuk konusu seçin:"
        ) if lang == 'tr' else (
            f"❓ Frequently Asked Questions:\n"
            f"📚 Please select a legal topic:"
        )
        reply_markup = get_faq_menu(lang)
    elif query.data == 'terms':
        USER_APPOINTMENT_STATE[user_id] = None
        message = (
            f"📚 Hukuki Terimler Sözlüğü:\n"
            f"📝 Lütfen bir terim seçin:"
        ) if lang == 'tr' else (
            f"📚 Legal Terms Dictionary:\n"
            f"📝 Please select a term:"
        )
        reply_markup = get_terms_menu(lang)
    elif query.data == 'tips':
        USER_APPOINTMENT_STATE[user_id] = None
        message = (
            f"💡 Hukuki İpuçları:\n"
            f"📋 Lütfen bir ipucu seçin:"
        ) if lang == 'tr' else (
            f"💡 Legal Tips:\n"
            f"📋 Please select a tip:"
        )
        reply_markup = get_tips_menu(lang)
    elif query.data == 'appointment':
        USER_APPOINTMENT_STATE[user_id] = 'select_date'
        message = (
            f"📅 Randevu Talebi:\n"
            f"🗓️ Lütfen bir tarih seçin:"
        ) if lang == 'tr' else (
            f"📅 Appointment Request:\n"
            f"🗓️ Please select a date:"
        )
        reply_markup = get_date_menu(lang)
    elif query.data.startswith('date_'):
        date = query.data.split('_')[1]
        USER_APPOINTMENT_STATE[user_id] = {'date': date, 'state': 'select_time'}
        message = (
            f"📅 Seçilen Tarih: {date}\n"
            f"🕒 Lütfen bir saat seçin:"
        ) if lang == 'tr' else (
            f"📅 Selected Date: {date}\n"
            f"🕒 Please select a time:"
        )
        reply_markup = get_time_menu(date, lang)
    elif query.data.startswith('time_'):
        _, date, time = query.data.split('_')
        USER_APPOINTMENT_STATE[user_id] = {'date': date, 'time': time, 'state': 'waiting_for_name'}
        message = (
            f"📅 Tarih: {date}, Saat: {time}\n"
            f"🖊️ Lütfen adınızı ve soyadınızı yazın (örneğin: Ahmet Yılmaz):"
        ) if lang == 'tr' else (
            f"📅 Date: {date}, Time: {time}\n"
            f"🖊️ Please enter your full name (e.g., Ahmet Yılmaz):"
        )
        reply_markup = get_back_button(lang)
    elif query.data.startswith('faq_'):
        USER_APPOINTMENT_STATE[user_id] = None
        faq_key = query.data.split('_', 1)[1]
        if faq_key in FAQ_DATA:
            message = FAQ_DATA[faq_key][lang].format(CURRENT_TIME)
        else:
            message = (
                f"😕 Bu SSS konusu bulunamadı.\n"
                f"📅 Son güncelleme: {CURRENT_TIME}"
            ) if lang == 'tr' else (
                f"😕 This FAQ topic was not found.\n"
                f"📅 Last updated: {CURRENT_TIME}"
            )
        reply_markup = get_back_button(lang)
    elif query.data.startswith('terms_'):
        USER_APPOINTMENT_STATE[user_id] = None
        term_key = query.data.split('_', 1)[1]
        if term_key in LEGAL_TERMS:
            message = LEGAL_TERMS[term_key][lang].format(CURRENT_TIME)
        else:
            message = (
                f"😕 Bu hukuki terim bulunamadı.\n"
                f"📅 Son güncelleme: {CURRENT_TIME}"
            ) if lang == 'tr' else (
                f"😕 This legal term was not found.\n"
                f"📅 Last updated: {CURRENT_TIME}"
            )
        reply_markup = get_back_button(lang)
    elif query.data.startswith('tips_'):
        USER_APPOINTMENT_STATE[user_id] = None
        tip_key = query.data.split('_', 1)[1]
        if tip_key in LEGAL_TIPS:
            message = LEGAL_TIPS[tip_key][lang].format(CURRENT_TIME)
        else:
            message = (
                f"😕 Bu hukuki ipucu bulunamadı.\n"
                f"📅 Son güncelleme: {CURRENT_TIME}"
            ) if lang == 'tr' else (
                f"😕 This legal tip was not found.\n"
                f"📅 Last updated: {CURRENT_TIME}"
            )
        reply_markup = get_back_button(lang)
    elif query.data == 'language':
        USER_APPOINTMENT_STATE[user_id] = None
        message = (
            f"🌐 Lütfen bir dil seçin:"
        ) if lang == 'tr' else (
            f"🌐 Please select a language:"
        )
        keyboard = [
            [InlineKeyboardButton("🇹🇷 Türkçe", callback_data='lang_tr')],
            [InlineKeyboardButton("🇬🇧 English", callback_data='lang_en')],
            [InlineKeyboardButton("⬅️ Geri" if lang == 'tr' else "⬅️ Back", callback_data='back')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
    elif query.data == 'feedback':
        USER_APPOINTMENT_STATE[user_id] = None
        message = (
            f"📝 Geri bildirim bırakmak için lütfen önerilerinizi yazın ve gönderin.\n"
            f"💡 Örnek: 'Bot harika, ama daha fazla hukuki bilgi eklenebilir!'\n"
            f"📅 Son güncelleme: {CURRENT_TIME}"
        ) if lang == 'tr' else (
            f"📝 Please write and send your feedback.\n"
            f"💡 Example: 'The bot is great, but more legal info could be added!'\n"
            f"📅 Last updated: {CURRENT_TIME}"
        )
        reply_markup = get_back_button(lang)
    elif query.data.startswith('lang_'):
        USER_APPOINTMENT_STATE[user_id] = None
        lang_code = query.data.split('_')[1]
        USER_LANGUAGE[user_id] = lang_code
        message = (
            f"🇹🇷 Dil Türkçe olarak ayarlandı! ⚖️\n"
            f"📅 Son güncelleme: {CURRENT_TIME}"
        ) if lang_code == 'tr' else (
            f"🇬🇧 Language set to English! ⚖️\n"
            f"📅 Last updated: {CURRENT_TIME}"
        )
        reply_markup = get_main_menu(lang_code)
    else:
        USER_APPOINTMENT_STATE[user_id] = None
        message = (
            f"😕 Geçersiz seçenek, lütfen tekrar deneyin.\n"
            f"📅 Son güncelleme: {CURRENT_TIME}"
        ) if lang == 'tr' else (
            f"😕 Invalid option, please try again.\n"
            f"📅 Last updated: {CURRENT_TIME}"
        )
        reply_markup = get_back_button(lang)

    # Mesajın değişip değişmediğini kontrol et
    if current_text != message or str(current_reply_markup) != str(reply_markup):
        await query.message.edit_text(message, reply_markup=reply_markup)
        USER_LAST_MESSAGE[user_id] = query.message.message_id

# Geri bildirim ve randevu mesajlarını işleme
async def handle_feedback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    lang = USER_LANGUAGE.get(user_id, 'tr')
    username = update.effective_user.username or "Bilinmiyor"
    user_input = update.message.text

    # Randevu alma süreci
    state = USER_APPOINTMENT_STATE.get(user_id, {})
    if isinstance(state, dict) and state.get('state') == 'waiting_for_name':
        USER_APPOINTMENT_STATE[user_id]['name'] = user_input
        USER_APPOINTMENT_STATE[user_id]['state'] = 'waiting_for_phone'
        message = (
            f"🖊️ Adınız kaydedildi: {user_input}\n"
            f"📱 Lütfen telefon numaranızı yazın (örneğin: +905551234567):"
        ) if lang == 'tr' else (
            f"🖊️ Name saved: {user_input}\n"
            f"📱 Please enter your phone number (e.g., +905551234567):"
        )
        reply_markup = get_back_button(lang)
    elif isinstance(state, dict) and state.get('state') == 'waiting_for_phone':
        if not is_valid_phone(user_input):
            message = (
                f"⚠️ Lütfen geçerli bir telefon numarası girin (örneğin: +905551234567):\n"
                f"📅 Son güncelleme: {CURRENT_TIME}"
            ) if lang == 'tr' else (
                f"⚠️ Please enter a valid phone number (e.g., +905551234567):\n"
                f"📅 Last updated: {CURRENT_TIME}"
            )
            reply_markup = get_back_button(lang)
        else:
            USER_APPOINTMENT_STATE[user_id]['phone'] = user_input
            USER_APPOINTMENT_STATE[user_id]['state'] = 'waiting_for_details'
            message = (
                f"📱 Telefon numarası kaydedildi: {user_input}\n"
                f"🗒️ Lütfen randevu detaylarını yazın (örneğin: Boşanma davası danışmanlığı):"
            ) if lang == 'tr' else (
                f"📱 Phone number saved: {user_input}\n"
                f"🗒️ Please enter the appointment details (e.g., Divorce case consultation):"
            )
            reply_markup = get_back_button(lang)
    elif isinstance(state, dict) and state.get('state') == 'waiting_for_details':
        name = USER_APPOINTMENT_STATE[user_id]['name']
        phone = USER_APPOINTMENT_STATE[user_id]['phone']
        date = USER_APPOINTMENT_STATE[user_id]['date']
        time = USER_APPOINTMENT_STATE[user_id]['time']
        details = user_input
        save_appointment(user_id, username, name, phone, date, time, details)
        USER_APPOINTMENT_STATE[user_id] = None
        message = (
            f"✅ Randevu talebiniz alındı! 🎉\n"
            f"🖊️ Ad: {name}\n"
            f"📱 Telefon: {phone}\n"
            f"📅 Tarih: {date}\n"
            f"🕒 Saat: {time}\n"
            f"🗒️ Detaylar: {details}\n"
            f"📩 Talebiniz kaydedildi, en kısa sürede sizinle iletişime geçeceğiz!\n"
            f"📅 Son güncelleme: {CURRENT_TIME}"
        ) if lang == 'tr' else (
            f"✅ Appointment request received! 🎉\n"
            f"🖊️ Name: {name}\n"
            f"📱 Phone: {phone}\n"
            f"📅 Date: {date}\n"
            f"🕒 Time: {time}\n"
            f"🗒️ Details: {details}\n"
            f"📩 Your request has been recorded, we will contact you soon!\n"
            f"📅 Last updated: {CURRENT_TIME}"
        )
        reply_markup = get_back_button(lang)
    else:
        # Geri bildirim işleme
        save_feedback(user_id, username, user_input)
        message = (
            f"🎉 Teşekkürler @{username}! Geri bildiriminiz alındı: '{user_input}' ⚖️\n"
            f"📅 Son güncelleme: {CURRENT_TIME}"
        ) if lang == 'tr' else (
            f"🎉 Thank you @{username}! Your feedback was received: '{user_input}' ⚖️\n"
            f"📅 Last updated: {CURRENT_TIME}"
        )
        reply_markup = get_back_button(lang)

    sent_message = await update.message.reply_text(message, reply_markup=reply_markup)
    USER_LAST_MESSAGE[user_id] = sent_message.message_id

def main():
    # Bot uygulamasını başlat
    app = Application.builder().token(TOKEN).build()

    # Komut ve buton işleyicilerini ekle
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CallbackQueryHandler(button))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_feedback))
    app.add_handler(MessageHandler(filters.COMMAND, unknown_command))

    # Botu çalıştır
    app.run_polling()

if __name__ == '__main__':
    main()
